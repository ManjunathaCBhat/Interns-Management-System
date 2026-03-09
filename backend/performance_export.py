"""
performance_export.py

"""

import base64
import io
import os
from datetime import datetime
from typing import Any, List, Optional

import httpx
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import JSONResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from pydantic import BaseModel

router = APIRouter(prefix="/admin/performance", tags=["Performance Export"])

# ─────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────

FROM_EMAIL = "interns360@cirruslabs.io"
FROM_NAME  = "Interns360 — CirrusLabs"


# ─────────────────────────────────────────────
# JWT — get logged-in user's email from token
# ─────────────────────────────────────────────

security = HTTPBearer()

def get_current_user_email(
    credentials: HTTPAuthorizationCredentials = Depends(security),
) -> str:
    """"
    Decodes the Bearer JWT and returns the logged-in user's email.
    Uses the same SECRET_KEY your login route uses.
    
    Your login route puts "sub" = username in the token, so we also
    look up the email from the DB — but since we don't have DB access
    here easily, we accept either "email" or "sub" from the token.
    
    If your token only has "sub" = username (not email), pass the email
    explicitly in the request body instead — see ExportEmailRequest below.
    """""
    token  = credentials.credentials
    secret = os.getenv("SECRET_KEY", "")

    if not secret:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="SECRET_KEY is not configured on the server.",
        )

    try:
        payload = jwt.decode(token, secret, algorithms=["HS256"])
        # Your login route sets: {"sub": username, "role": role}
        # It does NOT set email in the token — so we use the
        # "to_email" field from the request body as fallback (see route below)
        email: Optional[str] = payload.get("email") or payload.get("sub")
        if not email:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Token does not contain a usable identity field.",
            )
        return email
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token.",
        )


# ─────────────────────────────────────────────
# Request schema
# ─────────────────────────────────────────────

class ExportRow(BaseModel):
    Name: str
    Email: str
    Role: str
    Organization: Optional[str] = ""
    EmployeeID: Optional[str] = ""
    Batch: Optional[str] = ""
    InternType: Optional[str] = ""
    Project: Optional[str] = ""
    Phone: Optional[str] = ""
    College: Optional[str] = ""
    CGPA: Optional[Any] = ""
    JoinedDate: Optional[str] = ""
    TaskCount: Optional[int] = 0
    CompletedTasks: Optional[int] = 0
    DSUStreak: Optional[int] = 0
    Skills: Optional[str] = ""


class ExportEmailRequest(BaseModel):
    exportType: str                    # "all" | "batch" | "individual"
    exportBatch: Optional[str] = None
    userId: Optional[str] = None
    generatedAt: Optional[str] = None
    rows: List[ExportRow]
    """
    to_email is included as a fallback because your JWT only has
    "sub" = username, not the actual email address.
    The frontend should pass the logged-in user's email here.
    """
    to_email: Optional[str] = None


# ─────────────────────────────────────────────
# Microsoft Graph — reuses your existing credentials
# ─────────────────────────────────────────────

async def _get_graph_token() -> str:
    """Get Microsoft Graph access token — same logic as main.py"""
    tenant_id     = os.getenv("tenant_id")
    client_id     = os.getenv("client_id")
    client_secret = os.getenv("AZURE_SECRET_KEY")

    missing = [k for k, v in {
        "tenant_id": tenant_id,
        "client_id": client_id,
        "AZURE_SECRET_KEY": client_secret,
    }.items() if not v]

    if missing:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Microsoft Graph not configured. Missing: {', '.join(missing)}",
        )

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            token_url,
            data={
                "client_id":     client_id,
                "client_secret": client_secret,
                "scope":         "https://graph.microsoft.com/.default",
                "grant_type":    "client_credentials",
            },
            timeout=30.0,
        )

    if resp.status_code != 200:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Failed to get Graph token: {resp.text}",
        )

    token = resp.json().get("access_token")
    if not token:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Graph token response missing access_token.",
        )
    return token


async def _send_via_graph(
    to_email: str,
    excel_bytes: bytes,
    filename: str,
    export_type: str,
    generated_at: str,
) -> None:
    """
    Send email with Excel attachment via Microsoft Graph.
    Sends FROM the SENDER_MAIL account (same as password-reset emails).
    """
    sender = os.getenv("SENDER_MAIL")
    if not sender:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="SENDER_MAIL is not configured in environment variables.",
        )

    token = await _get_graph_token()

    # Base64-encode the Excel file for the Graph attachment payload
    excel_b64 = base64.b64encode(excel_bytes).decode("utf-8")

    body_html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333;">
        <div style="max-width:600px; margin:auto; padding:24px;
                    border:1px solid #e0e0e0; border-radius:8px;">

            <div style="background: linear-gradient(to right, #0F0E47, #505081);
                        padding:16px 24px; border-radius:6px 6px 0 0;">
                <h2 style="color:white; margin:0;">📊 Performance Report</h2>
            </div>

            <div style="padding:24px;">
                <p>Hi,</p>
                <p>Your requested performance report is attached to this email.</p>

                <table style="border-collapse:collapse; width:100%; margin:16px 0;">
                    <tr style="background:#f5f5f5;">
                        <td style="padding:8px 12px; font-weight:bold;
                                   border:1px solid #ddd; width:140px;">Export Type</td>
                        <td style="padding:8px 12px; border:1px solid #ddd;">{export_type.title()}</td>
                    </tr>
                    <tr>
                        <td style="padding:8px 12px; font-weight:bold; border:1px solid #ddd;">File</td>
                        <td style="padding:8px 12px; border:1px solid #ddd;">{filename}</td>
                    </tr>
                    <tr style="background:#f5f5f5;">
                        <td style="padding:8px 12px; font-weight:bold; border:1px solid #ddd;">Generated On</td>
                        <td style="padding:8px 12px; border:1px solid #ddd;">{generated_at[:10]}</td>
                    </tr>
                    <tr>
                        <td style="padding:8px 12px; font-weight:bold; border:1px solid #ddd;">Sent To</td>
                        <td style="padding:8px 12px; border:1px solid #ddd;">{to_email}</td>
                    </tr>
                </table>

                <p style="color:#888; font-size:12px; margin-top:24px;">
                    This email was sent automatically from the CirrusLabs Performance Dashboard.<br/>
                    Please do not reply to this email.
                </p>
            </div>
        </div>
    </body>
    </html>
    """

    graph_payload = {
        "message": {
            "subject": f"Performance Report — {export_type.title()} ({generated_at[:10]})",
            "body": {
                "contentType": "HTML",
                "content": body_html,
            },
            "toRecipients": [
                {"emailAddress": {"address": to_email}}
            ],
            "attachments": [
                {
                    "@odata.type":  "#microsoft.graph.fileAttachment",
                    "name":         filename,
                    "contentType":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "contentBytes": excel_b64,
                }
            ],
        },
        "saveToSentItems": "false",
    }

    send_url = f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail"
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            send_url,
            json=graph_payload,
            headers={"Authorization": f"Bearer {token}"},
            timeout=30.0,
        )

    if resp.status_code not in (200, 202):
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Graph sendMail failed ({resp.status_code}): {resp.text}",
        )


# ─────────────────────────────────────────────
# Build Excel in memory
# ─────────────────────────────────────────────

def _build_excel_bytes(rows: List[ExportRow], generated_at: str) -> bytes:
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    headers = [
        "Name", "Email", "Role", "Organization", "EmployeeID",
        "Batch", "InternType", "Project", "Phone", "College",
        "CGPA", "JoinedDate", "TaskCount", "CompletedTasks",
        "DSUStreak", "Skills",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0F0E47", end_color="0F0E47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([
            row.Name, row.Email, row.Role, row.Organization,
            row.EmployeeID, row.Batch, row.InternType, row.Project,
            row.Phone, row.College, row.CGPA, row.JoinedDate,
            row.TaskCount, row.CompletedTasks, row.DSUStreak, row.Skills,
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    meta = wb.create_sheet("Info")
    meta.append(["Generated At", generated_at])
    meta.append(["Total Records", len(rows)])
    meta.append(["Sent From", FROM_EMAIL])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# Route
# ─────────────────────────────────────────────

@router.post("/export-email", status_code=status.HTTP_200_OK)
async def export_performance_email(
    payload: ExportEmailRequest,
    token_identity: str = Depends(get_current_user_email),
):
    """
    FROM : interns360@cirruslabs.io  (sent via Microsoft Graph / SENDER_MAIL)
    TO   : logged-in user's email from payload.to_email
           (fallback: "sub" from JWT if it happens to be an email)

    Uses the same tenant_id / client_id / AZURE_SECRET_KEY already in .env.
    """

    # ── Resolve TO email ─────────────────────────────────────────────────
    # Your JWT has "sub" = username (not email), so we rely on the
    # frontend passing to_email in the body.
    to_email = payload.to_email or (token_identity if "@" in token_identity else None)
    if not to_email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not determine recipient email. Pass 'to_email' in the request body.",
        )

    # ── Empty data guard ─────────────────────────────────────────────────
    if not payload.rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data to export. The selected filter returned 0 records.",
        )

    # ── Build filename ────────────────────────────────────────────────────
    date_str     = datetime.utcnow().strftime("%Y-%m-%d")
    generated_at = payload.generatedAt or datetime.utcnow().isoformat()

    if payload.exportType == "batch" and payload.exportBatch:
        filename = f"performance_batch_{payload.exportBatch}_{date_str}.xlsx"
    elif payload.exportType == "individual" and payload.rows:
        name     = payload.rows[0].Name.replace(" ", "_")
        filename = f"performance_{name}_{date_str}.xlsx"
    else:
        filename = f"performance_all_{date_str}.xlsx"

    # ── Generate Excel ────────────────────────────────────────────────────
    try:
        excel_bytes = _build_excel_bytes(payload.rows, generated_at)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate Excel file: {str(exc)}",
        )

    # ── Send via Microsoft Graph ──────────────────────────────────────────
    await _send_via_graph(
        to_email=to_email,
        excel_bytes=excel_bytes,
        filename=filename,
        export_type=payload.exportType,
        generated_at=generated_at,
    )

    return JSONResponse(
        status_code=200,
        content={
            "success":  True,
            "message":  f"Report sent to {to_email}",
            "from":     os.getenv("SENDER_MAIL", FROM_EMAIL),
            "to":       to_email,
            "filename": filename,
            "records":  len(payload.rows),
        },
    )